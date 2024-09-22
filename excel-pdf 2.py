from openpyxl import load_workbook
from reportlab.pdfgen import canvas

def excel_to_pdf(excel_file_path, pdf_file_path):
    wb = load_workbook(filename=excel_file_path)
    sheet = wb.active

    pdf = canvas.Canvas(pdf_file_path)
    row_height = 20
    x_offset = 50
    y_offset = 800

    for row in sheet.iter_rows(values_only=True):
        y_offset -= row_height
        for value in row:
            pdf.drawString(x_offset, y_offset, str(value))
            x_offset += 100
        x_offset = 50

    pdf.save()

# Path to the Excel file
excel_file_path = "path_to_your_excel_file.xlsx"

# Path to save the PDF file
pdf_file_path = "path_to_save_pdf_file.pdf"

# Convert the Excel file to PDF
excel_to_pdf(excel_file_path, pdf_file_path)
